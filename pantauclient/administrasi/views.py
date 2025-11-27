import re
from multiprocessing import context
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse, HttpResponseBadRequest
from django.contrib.auth import authenticate, login
from django.contrib import messages
from akademik.models import Kelas, Siswa, Mapel, JamPelajaran, Jadwal, UserGuru, ProfileGuru, GuruMapel, PresensiSekolah, PresensiStatus
from akademik.forms import KelasForm, MapelForm, JamPelajaranForm, JadwalForm, UserGuruForm, ProfileGuruForm
from django.db import transaction
from django.contrib.auth.hashers import make_password
from django.contrib import messages
from django.db.models import Value, CharField, Q, Count, IntegerField, When, Case
from datetime import timedelta
from django.db.models.expressions import Case, When
import csv, io, re
from django.utils.dateparse import parse_date
import json
from django.views.decorators.http import require_POST
from django.utils import timezone
from django.forms.models import model_to_dict
import openpyxl
from openpyxl.utils import get_column_letter
from .wa_bot import generate_laporan_excel, kirim_laporan_wa_otomatis
import threading # Agar browser jalan di background tidak bikin loading lama

from .wa_bot import proses_laporan_wa


# Create your views here.
def index(request):
    context = {
        'title': 'Dashboard',
        # AKTIFASI TOMBOL
        'icon': 'bg-[#3d3d3d]',
    }
    return render(request, 'administrasi/index.html', context)

def rekap(request):
    kelas = Kelas.objects.all()
    context = {
        'semua_kelas': kelas,
        'title': 'Rekap',
    }
    return render(request, 'administrasi/rekap.html', context)

def pengaturan(request):
    context = {
        'title': 'Pengaturan',
    }
    return render(request, 'administrasi/pengaturan.html', context)

def guru(request):
    semua_guru = UserGuru.objects.all()
    context = {
        'title': 'Manajemen Data Guru',
        'semua_objek': semua_guru,
        'page': 'guru',
    }
    return render(request, 'administrasi/guru/guru.html', context)

def kelas(request):
    semua_kelas = Kelas.objects.all()
    context = {
        'title': 'Manajemen Jadwal dan Kelas',
        'semua_objek': semua_kelas,
        'page': 'kelas',
    }
    return render(request, 'administrasi/jadwalkelas/kelas.html', context)

def jampel(request):
    semua_jampel = JamPelajaran.objects.all()
    context = {
        'title': 'Manajemen Jadwal dan Kelas',
        'semua_objek': semua_jampel,
        'page': 'jampel',
    }
    return render(request, 'administrasi/jadwalkelas/jampel.html', context)

def mapel(request):
    semua_mapel = Mapel.objects.all()
    context = {
        'title': 'Manajemen Jadwal dan Kelas',
        'semua_objek': semua_mapel,
        'page': 'mapel',
    }
    return render(request, 'administrasi/jadwalkelas/mapel.html', context)

def jadwal(request):
    semua_jadwal = Jadwal.objects.all()
    context = {
        'title': 'Manajemen Jadwal dan Kelas',
        'semua_objek': semua_jadwal,
        'page': 'jadwal',
    }
    return render(request, 'administrasi/jadwalkelas/jadwal.html', context)

def get_guru(request):
    mapel_id = request.GET.get('mapel_id')
    gurus = UserGuru.objects.filter(mengajar__id=mapel_id).values('id', 'nama')
    return JsonResponse(list(gurus), safe=False)

def jadwal_edit(request, page, obj_id=None):
    """View untuk menangani tambah/edit data Kelas, Mapel, Jam Pelajaran, dan Jadwal."""
    model_map = {
        'kelas': (Kelas, KelasForm),
        'mapel': (Mapel, MapelForm),
        'jampel': (JamPelajaran, JamPelajaranForm),
        'jadwal': (Jadwal, JadwalForm),
    }
    
    Model, Form = model_map.get(page)
    obj = None
    if obj_id:
        obj = get_object_or_404(Model, id=obj_id)

    if request.method == 'POST':
        form = Form(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            return redirect(page)
    else:
        form = Form(instance=obj)

    context = {
        'title': 'Manajemen Jadwal dan Kelas',
        'form': form,
        'page': page,
        'obj': obj,
    }
    return render(request, 'administrasi/jadwalkelas/edit.html', context)

@transaction.atomic
def guru_edit(request, obj_id=None):
    """
    Admin:
    - TIDAK bisa ganti password lewat form.
    - CREATE: password = NIP (di-hash).
    - RESET PASSWORD: password = NIP (di-hash).
    """
    if obj_id:
        user_guru = get_object_or_404(UserGuru, id=obj_id)
        profile_guru = getattr(user_guru, 'profileguru', None) or ProfileGuru(guru=user_guru)
    else:
        user_guru = UserGuru()
        profile_guru = ProfileGuru()

    if request.method == 'POST':
        # Tombol reset password
        if 'reset_password' in request.POST and (obj_id or user_guru.id):
            guru_obj = user_guru if user_guru.id else get_object_or_404(UserGuru, id=obj_id)
            guru_obj.password = make_password(guru_obj.nip)
            guru_obj.save(update_fields=['password'])
            messages.success(request, 'Password berhasil di-reset ke NIP.')
            return redirect('edit_guru', obj_id=guru_obj.id)

        user_form = UserGuruForm(request.POST, instance=user_guru)
        profile_form = ProfileGuruForm(request.POST, instance=profile_guru)

        # === Penting: longgarkan requirement tanpa ubah tampilan ===
        if 'password' in user_form.fields:
            user_form.fields['password'].required = False
        if 'mengajar' in user_form.fields:
            user_form.fields['mengajar'].required = False  # karena tidak ada inputnya di template

        if user_form.is_valid() and profile_form.is_valid():
            is_create = user_guru.id is None
            user = user_form.save(commit=False)

            if is_create:
                # password awal mengikuti NIP
                user.password = make_password(user.nip)
            else:
                # Saat edit, jangan sentuh password
                original = UserGuru.objects.get(id=user_guru.id)
                user.password = original.password

            user.save()
            # relasi M2M (abaikan kalau tidak ada inputnya—tidak mengapa)
            try:
                user_form.save_m2m()
            except Exception:
                pass

            profile = profile_form.save(commit=False)
            profile.guru = user
            profile.save()

            messages.success(request, 'Data guru berhasil disimpan.')
            return redirect('guru')
        else:
            # Tampilkan error ke user (tanpa ubah UI lain)
            if user_form.errors:
                messages.error(request, f"User form error: {user_form.errors.as_text()}")
            if profile_form.errors:
                messages.error(request, f"Profile form error: {profile_form.errors.as_text()}")

    else:
        user_form = UserGuruForm(instance=user_guru)
        profile_form = ProfileGuruForm(instance=profile_guru)
        # Longgarkan juga di GET agar konsisten (tidak wajib diisi)
        if 'password' in user_form.fields:
            user_form.fields['password'].required = False
        if 'mengajar' in user_form.fields:
            user_form.fields['mengajar'].required = False
    
    list_gm = []
    if user_guru.id:
        list_gm = GuruMapel.objects.filter(guru=user_guru).select_related('mapel')

    context = {
        'title': 'Manajemen Data Guru',
        'user_form': user_form,
        'profile_form': profile_form,
        'page': 'guru',
        'obj': user_guru,
        'list_gurumapel': list_gm,
    }
    return render(request, 'administrasi/guru/guru_edit.html', context)

def hapus(request, page, obj_id):
    model_map = {
        'kelas': Kelas,
        'mapel': Mapel,
        'jampel': JamPelajaran,
        'jadwal': Jadwal,
        'guru': UserGuru,  # <-- TAMBAHKAN INI
    }

    Model = model_map.get(page)
    if not Model:
        messages.error(request, "Halaman hapus tidak dikenal.")
        return redirect('index')

    obj = get_object_or_404(Model, id=obj_id)
    nama_obj = getattr(obj, 'nama', str(obj))  # biar ada nama saat pesan
    obj.delete()

    # Redirect sesuai halaman
    redirect_map = {
        'kelas': 'kelas',
        'mapel': 'mapel',
        'jampel': 'jampel',
        'jadwal': 'jadwal',
        'guru': 'guru',
    }
    messages.success(request, f"Data {page} '{nama_obj}' berhasil dihapus.")
    return redirect(redirect_map.get(page, 'index'))

@transaction.atomic
def guru_mapel_tambah(request, guru_id):
    guru = get_object_or_404(UserGuru, id=guru_id)
    semua_mapel = Mapel.objects.all().order_by('nama_mapel')

    if request.method == 'POST':
        mapel_id = request.POST.get('mapel')
        if not mapel_id:
            messages.error(request, 'Silakan pilih mata pelajaran.')
            return redirect('tambah_mapel_guru', guru_id=guru.id)

        mapel = get_object_or_404(Mapel, id=mapel_id)

        # kode_ajar = kode_mapel + 3 digit terakhir dari NIP (hanya digit)
        nip_digits = re.sub(r'\D', '', guru.nip or '')
        suffix = nip_digits[-3:] if len(nip_digits) >= 3 else nip_digits
        kode_ajar = f"{mapel.kode_mapel}{suffix}"

        # Cegah duplikasi (guru-mapel yang sama)
        if GuruMapel.objects.filter(guru=guru, mapel=mapel).exists():
            messages.warning(request, 'Guru sudah mengampu mapel tersebut.')
            return redirect('edit_guru', obj_id=guru.id)

        gm = GuruMapel.objects.create(
            guru=guru,
            mapel=mapel,
            kode_ajar=kode_ajar
        )
        messages.success(request, f"Mapel '{mapel.nama_mapel}' ditambahkan. Kode ajar: {gm.kode_ajar}")
        return redirect('edit_guru', obj_id=guru.id)

    # GET → tampilkan form pilih mapel
    context = {
        'title': 'Tambah Mapel Diampu',
        'guru': guru,
        'semua_mapel': semua_mapel,
    }
    return render(request, 'administrasi/guru/guru_mapel_tambah.html', context)

@transaction.atomic
def guru_mapel_hapus(request, guru_id, gm_id):
    guru = get_object_or_404(UserGuru, id=guru_id)
    gm = get_object_or_404(GuruMapel, id=gm_id, guru=guru)
    gm.delete()
    messages.success(request, 'Mapel berhasil dihapus dari daftar ampuan.')
    return redirect('edit_guru', obj_id=guru.id)


# =========================
#  HALAMAN DAFTAR SISWA
# =========================
try:
    from openpyxl import load_workbook  # untuk xlsx
except Exception:
    load_workbook = None


def siswa(request):
    semua_kelas = Kelas.objects.all().order_by('nama_kelas')
    context = {
        'title': 'Siswa',
        'semua_kelas': semua_kelas,
    }
    return render(request, 'administrasi/siswa/siswa.html', context)

from django.http import JsonResponse

def get_siswa(request):
    """Endpoint AJAX untuk pencarian & filter siswa (instan)."""
    q = request.GET.get('q', '').strip()
    kelas_id = request.GET.get('kelas_id')

    # --- PERBAIKAN DI SINI ---
    # Gunakan select_related untuk mengambil data kelas dalam satu query
    siswa_qs = Siswa.objects.select_related('kelas').all().order_by('nama')
    # -------------------------

    if kelas_id:
        try:
            siswa_qs = siswa_qs.filter(kelas_id=int(kelas_id))
        except ValueError:
            pass
    if q:
        from django.db.models import Q
        siswa_qs = siswa_qs.filter(
            Q(nama__icontains=q) | Q(nisn__icontains=q) | Q(nipd__icontains=q)
        )

    def wali_of(s):
        return s.nama_wali or (s.nama_ayah or s.nama_ibu) or ''

    data = [{
        'id': s.id,
        'nama': s.nama,
        'nisn': s.nisn,
        'nipd': s.nipd or '',
        'kelas': s.kelas.nama_kelas, # Baris ini sekarang tidak memicu query baru
        'wali': wali_of(s),
        'no_siswa': s.no_siswa or '',
        'no_ortu': s.no_ortu or '',
    } for s in siswa_qs[:500]]

    return JsonResponse({'items': data})

@transaction.atomic
def siswa_import(request):
    """
    Flow:
    - GET: tampilkan form kosong
    - POST (preview): baca file, validasi per baris, simpan hasil preview ke session, render preview
    - POST (save): ambil preview dari session; jika masih ada error -> tolak; jika bersih -> simpan ke DB
    """
    SESSION_KEY = "siswa_import_preview"

    def normalize_header(h):
        return (h or '').strip().lower().replace(' ', '_')

    def get(r, *keys):
        for k in keys:
            if k in r and r[k] != '':
                return r[k]
        return ''

    # === SAVE (tanpa file) → gunakan data dari session ===
    if request.method == 'POST' and request.POST.get('action') == 'save' and not request.FILES.get('file'):
        preview_rows = request.session.get(SESSION_KEY, [])
        if not preview_rows:
            messages.error(request, "Sesi preview kosong atau kadaluarsa. Silakan upload file lagi.")
            return redirect('siswa_import')

        has_row_errors = any(r.get('errors') for r in preview_rows)
        if has_row_errors:
            messages.error(request, "Tidak dapat menyimpan: masih ada baris bermasalah (ditandai merah). Perbaiki dan upload ulang.")
            context = {'title': 'Import Siswa', 'preview_rows': preview_rows, 'errors': [], 'has_row_errors': True}
            return render(request, 'administrasi/siswa_import.html', context)

        # Simpan ke DB
        saved = 0
        for r in preview_rows:
            kelas_obj = Kelas.objects.get(nama_kelas=r['kelas'])  # sudah divalidasi ada
            jk_val = (r['jk'] or '').upper()
            jk_db = 'P' if jk_val in ('P', 'PEREMPUAN') else 'L'
            tgl_obj = parse_date(r['tanggal_lahir'])
            Siswa.objects.update_or_create(
                nisn=r['nisn'],
                defaults={
                    'nama': r['nama'],
                    'nipd': r['nipd'] or None,
                    'jenis_kelamin': jk_db,
                    'tempat_lahir': r['tempat_lahir'] or '',
                    'tanggal_lahir': tgl_obj,
                    'nama_ayah': r['nama_ayah'] or None,
                    'nama_ibu': r['nama_ibu'] or None,
                    'nama_wali': r['nama_wali'] or None,
                    'kelas': kelas_obj,
                    'no_siswa': r['no_siswa'] or None,
                    'no_ortu': r['no_ortu'] or None,
                }
            )
            saved += 1

        # bersihkan session setelah simpan
        try:
            del request.session[SESSION_KEY]
        except KeyError:
            pass

        messages.success(request, f'Berhasil menyimpan {saved} data siswa.')
        return redirect('siswa')

    # === PREVIEW (upload file) ===
    if request.method == 'POST' and request.FILES.get('file'):
        f = request.FILES['file']
        name = f.name.lower()
        rows = []
        errors_global = []
        preview_rows = []

        # Baca file (CSV / Excel)
        try:
            if name.endswith('.csv'):
                data = io.TextIOWrapper(f.file, encoding='utf-8', errors='ignore')
                reader = csv.DictReader(data)
                for r in reader:
                    rows.append({normalize_header(k): (v or '').strip() for k, v in r.items()})
            else:
                from openpyxl import load_workbook
                wb = load_workbook(f, data_only=True)
                ws = wb.active
                headers = [normalize_header(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))[0:ws.max_column]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    r = {}
                    for idx, val in enumerate(row):
                        key = headers[idx] if idx < len(headers) else f'col_{idx}'
                        r[key] = (str(val).strip() if val is not None else '')
                    rows.append(r)
        except Exception as e:
            errors_global.append(f'Gagal membaca file: {e}')

        # Normalisasi + validasi per baris
        for ix, r in enumerate(rows, start=2):
            nama = get(r, 'nama', 'nama_siswa')
            nisn = get(r, 'nisn')
            nipd = get(r, 'nipd', 'nis')
            kelas_nama = get(r, 'kelas', 'nama_kelas', 'rombel', 'kelas_rombel')
            jk_raw = get(r, 'jk', 'jenis_kelamin')
            tempat = get(r, 'tempat_lahir', 'tempat')
            tgl_raw = get(r, 'tanggal_lahir', 'tgl_lahir', 'tanggal')
            ayah = get(r, 'nama_ayah', 'ayah')
            ibu = get(r, 'nama_ibu', 'ibu')
            wali = get(r, 'nama_wali', 'wali')
            no_siswa = get(r, 'no_siswa', 'hp_siswa', 'No. Siswa')
            no_ortu = get(r, 'no_ortu', 'hp_ortu', 'No. Ortu')

            row_errors = []
            if not nama:
                row_errors.append("Nama wajib diisi")
            if not nisn:
                row_errors.append("NISN wajib diisi")
            if not kelas_nama:
                row_errors.append("Kelas wajib diisi")

            # kelas harus sudah ada (Kelas punya relasi OneToOne ke guru)
            kelas_obj = None
            if kelas_nama:
                kelas_obj = Kelas.objects.filter(nama_kelas=kelas_nama).first()
                if not kelas_obj:
                    row_errors.append(f"Kelas '{kelas_nama}' belum ada di database")

            # tanggal lahir wajib (model DateField tanpa null=True)
            tgl_obj = parse_date(tgl_raw) if tgl_raw else None
            if not tgl_obj:
                row_errors.append("Tanggal lahir wajib & format YYYY-MM-DD")

            preview_rows.append({
                'rownum': ix,
                'errors': row_errors,
                'nama': nama,
                'nisn': nisn,
                'nipd': nipd,
                'kelas': kelas_nama,
                'jk': jk_raw,
                'tempat_lahir': tempat,
                'tanggal_lahir': tgl_raw,
                'nama_ayah': ayah,
                'nama_ibu': ibu,
                'nama_wali': wali,
                'no_siswa': no_siswa,
                'no_ortu': no_ortu,
            })

        has_row_errors = any(r['errors'] for r in preview_rows)

        # Simpan ke session untuk dipakai saat klik Simpan
        request.session[SESSION_KEY] = preview_rows

        context = {
            'title': 'Import Siswa',
            'preview_rows': preview_rows,
            'errors': errors_global,
            'has_row_errors': has_row_errors,
        }
        return render(request, 'administrasi/siswa/siswa_import.html', context)

    # GET
    context = {'title': 'Import Siswa'}
    return render(request, 'administrasi/siswa/siswa_import.html', context)


# =========================
#   Halaman Presensi
# =========================
def deteksi(request):
    context = {
        'title': 'Presensi Deteksi Wajah',
    }
    return render(request, 'administrasi/presensi/deteksi.html', context)

def import_gambar(request):
    context = {
        'title': 'Presensi Scan Gambar',
    }
    return render(request, 'administrasi/presensi/import_gambar.html', context)

@require_POST
def deteksi_frame(request):
    """
    Terima 1 frame, deteksi wajah, dan KEMBALIKAN DATA LENGKAP (Termasuk Kelas).
    """
    data = json.loads(request.body.decode('utf-8'))
    img_b64 = data.get('image')
    if not img_b64:
        return HttpResponseBadRequest("image wajib diisi")

    from akademik.face_detect import detect_from_base64
    from akademik.models import Siswa # Pastikan import ini ada
    
    # 1. Deteksi Wajah
    detected = detect_from_base64(img_b64)
    
    # 2. Ambil Info Kelas dari Database (Optimasi Query)
    # Kumpulkan semua NISN yang terdeteksi
    nisn_list = [d.get('nisn') for d in detected if d.get('nisn')]
    
    # Mapping NISN -> Nama Kelas
    kelas_map = {}
    if nisn_list:
        siswas = Siswa.objects.filter(nisn__in=nisn_list).select_related('kelas')
        for s in siswas:
            kelas_map[s.nisn] = s.kelas.nama_kelas if s.kelas else "-"

    # 3. Gabungkan Data
    ts = timezone.localtime()
    ts_iso = ts.isoformat()
    
    for d in detected:
        d['timestamp'] = ts_iso
        # Masukkan info kelas ke respon JSON
        if d.get('nisn'):
            d['kelas'] = kelas_map.get(d['nisn'], '-')
        else:
            d['kelas'] = '-'

    return JsonResponse({'detected': detected})

@require_POST
def deteksi_resolve(request):
    """
    Body JSON: { "nisns": ["1234567890", ...] }
    Return: { items: [ {nisn, nama, kelas}, ... ] }
    """
    import json
    try:
        payload = json.loads(request.body.decode('utf-8'))
    except Exception:
        return HttpResponseBadRequest("Body JSON tidak valid")

    nisns = payload.get('nisns') or []
    if not isinstance(nisns, list) or not all(isinstance(x, str) for x in nisns):
        return HttpResponseBadRequest("nisns harus list of string")

    from akademik.models import Siswa
    qs = Siswa.objects.select_related('kelas').filter(nisn__in=nisns)
    items = []
    for s in qs:
        items.append({
            'nisn': s.nisn,
            'nama': s.nama,
            'kelas': getattr(getattr(s, 'kelas', None), 'nama_kelas', '-')
        })
    # untuk NISN yang belum dikenal di DB, tetap kembalikan placeholder
    known = set(x['nisn'] for x in items)
    for z in nisns:
        if z not in known:
            items.append({'nisn': z, 'nama': '(tidak dikenal)', 'kelas': '-'})
    return JsonResponse({'items': items})



# Tambahkan view ini di bawah view import_gambar
def manual(request):
    semua_kelas = Kelas.objects.all().order_by('nama_kelas')
    context = {
        'title': 'Presensi Manual',
        'semua_kelas': semua_kelas,
    }
    return render(request, 'administrasi/presensi/manual.html', context)

def get_siswa_presensi_manual(request):
    """
    Endpoint untuk mengambil daftar siswa per kelas beserta status presensi
    pada tanggal tertentu.
    """
    kelas_id = request.GET.get('kelas_id')
    tanggal_str = request.GET.get('tanggal')
    
    if not kelas_id or not tanggal_str:
        return JsonResponse({'error': 'Parameter kelas_id dan tanggal dibutuhkan'}, status=400)

    try:
        tanggal = timezone.datetime.strptime(tanggal_str, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({'error': 'Format tanggal tidak valid'}, status=400)

    from akademik.models import PresensiSekolah, PresensiStatus
    
    # Ambil semua siswa di kelas tersebut
    list_siswa = Siswa.objects.filter(kelas_id=kelas_id).order_by('nama')
    
    # Ambil data presensi yang sudah ada untuk siswa-siswa tersebut pada tanggal itu
    presensi_exist = PresensiSekolah.objects.filter(
        siswa__in=list_siswa,
        tanggal=tanggal
    ).select_related('siswa')
    
    status_map = {p.siswa.id: p.status for p in presensi_exist}
    
    data_siswa = []
    for s in list_siswa:
        inisial = "".join(part[0] for part in s.nama.split()[:2]).upper()
        data_siswa.append({
            'id': s.id,
            'nama': s.nama,
            'nisn': s.nisn,
            'status': status_map.get(s.id, PresensiStatus.H), # Default Hadir (H)
            'photoInitial': inisial,
        })

    return JsonResponse({'siswa': data_siswa})

# =============================================
#  LOGIKA 1: PRESENSI MANUAL (SIMPAN + WA)
# =============================================
@require_POST
@transaction.atomic
def save_presensi_manual(request):
    try:
        data = json.loads(request.body)
        tanggal_str = data.get('tanggal')
        updates = data.get('updates')
        
        if not tanggal_str or not updates:
            return JsonResponse({'error': 'Data tidak lengkap'}, status=400)

        tanggal = timezone.datetime.strptime(tanggal_str, '%Y-%m-%d').date()
        saved_count = 0
        
        for item in updates:
            siswa_id = item.get('id')
            status = item.get('status')
            
            siswa_obj = get_object_or_404(Siswa, id=siswa_id)
            waktu = timezone.make_aware(timezone.datetime.combine(tanggal, timezone.datetime.min.time()))

            PresensiSekolah.objects.update_or_create(
                siswa=siswa_obj,
                tanggal=tanggal,
                defaults={'status': status, 'waktu': waktu}
            )
            saved_count += 1
            
        # --- PERBAIKAN: TIDAK ADA PENGIRIMAN WA DI SINI ---
        # Laporan WA hanya dikirim lewat menu 'Tutup Presensi'
        
        return JsonResponse({
            'ok': True, 
            'message': f'{saved_count} data berhasil disimpan secara manual.'
        })

    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)

# =============================================
#  LOGIKA 2: DETEKSI WAJAH / SCAN (SIMPAN + WA)
# =============================================
@require_POST
@transaction.atomic
def deteksi_save(request):
    """
    Hanya simpan status HADIR dari deteksi wajah/scan.
    Tidak ada Alpha otomatis, Tidak ada WA otomatis.
    """
    try:
        payload = json.loads(request.body.decode('utf-8'))
        arr = payload.get('detected') or []
        
        if not isinstance(arr, list):
            return HttpResponseBadRequest("Format data salah")

        # Grouping per tanggal
        by_date = {}
        for it in arr:
            nisn = (it.get('nisn') or '').strip()
            ts = it.get('timestamp')
            if not nisn: continue
            
            try:
                ts_dt = timezone.datetime.fromisoformat(ts)
                if timezone.is_naive(ts_dt):
                    ts_dt = timezone.make_aware(ts_dt, timezone.get_current_timezone())
            except:
                ts_dt = timezone.localtime()
            
            d = ts_dt.date()
            by_date.setdefault(d, []).append((nisn, ts_dt))

        total_saved = 0
        
        for d, items in by_date.items():
            # Filter duplikat lokal
            seen = {}
            for nisn, ts_dt in items:
                seen.setdefault(nisn, ts_dt)
                if ts_dt < seen[nisn]: seen[nisn] = ts_dt

            # Ambil data siswa
            nisn_list = list(seen.keys())
            siswa_qs = Siswa.objects.filter(nisn__in=nisn_list)
            siswa_map = {s.nisn: s for s in siswa_qs}
            
            for nisn, ts_dt in seen.items():
                s = siswa_map.get(nisn)
                if s:
                    # Update jadi HADIR
                    PresensiSekolah.objects.update_or_create(
                        siswa=s, tanggal=d,
                        defaults={'status': PresensiStatus.H, 'waktu': ts_dt}
                    )
                    total_saved += 1

        # --- PERBAIKAN: TIDAK ADA PENGIRIMAN WA DI SINI ---

        return JsonResponse({'ok': True, 'saved': total_saved})

    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)

def tutup_presensi_harian(request):
    """
    Fungsi Pamungkas. Dijalankan manual oleh Admin setelah jam masuk.
    1. Cari siswa yang belum punya status hari ini -> Tandai ALPHA.
    2. Kirim Laporan WA ke SEMUA KELAS yang ada siswanya.
    """
    if request.method != 'POST':
        return HttpResponseBadRequest("Harus POST")

    tanggal = timezone.localtime().date() # Hari ini
    waktu_alpha = timezone.make_aware(timezone.datetime.combine(tanggal, timezone.datetime.min.time()))
    
    # 1. TANDAI ALPHA MASSAL
    # Ambil semua ID siswa aktif
    all_siswa = Siswa.objects.all().values('id', 'kelas_id')
    
    # Ambil ID siswa yang SUDAH absen (Hadir/Izin/Sakit/Terlambat) hari ini
    sudah_absen_ids = set(
        PresensiSekolah.objects.filter(tanggal=tanggal).values_list('siswa_id', flat=True)
    )
    
    bulk_alpha = []
    kelas_terdampak_ids = set()
    
    for s in all_siswa:
        if s['id'] not in sudah_absen_ids:
            # Belum absen -> Alpha
            bulk_alpha.append(PresensiSekolah(
                siswa_id=s['id'],
                tanggal=tanggal,
                status=PresensiStatus.A,
                waktu=waktu_alpha
            ))
            if s['kelas_id']:
                kelas_terdampak_ids.add(s['kelas_id'])
        else:
            # Sudah absen -> Masukkan kelasnya ke daftar kirim WA juga
            # (Agar kelas yang hadir semua tetap dapat laporan)
            if s['kelas_id']:
                kelas_terdampak_ids.add(s['kelas_id'])
    
    # Simpan Alpha ke Database
    if bulk_alpha:
        PresensiSekolah.objects.bulk_create(bulk_alpha, ignore_conflicts=True)
    
    # 2. KIRIM WA MASSAL (ANTREAN)
    # Import helper kita
    from .wa_bot import proses_laporan_wa
    
    jumlah_laporan = 0
    for k_id in kelas_terdampak_ids:
        sukses = proses_laporan_wa(k_id, tanggal, mode="SCAN")
        if sukses:
            jumlah_laporan += 1
            
    messages.success(request, f"Presensi Ditutup! {len(bulk_alpha)} siswa ditandai Alpha. {jumlah_laporan} laporan WA masuk antrean.")
    return redirect('rekap') # Atau redirect ke dashboard

# GANTI FUNGSI LAMA DENGAN YANG INI
def get_rekap_data(request):
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    # Ambil filter dari frontend
    scope = request.GET.get('scope') # 'semua', 'tingkatan', 'kelas'
    scope_value = request.GET.get('scope_value') # ID tingkatan atau ID kelas

    if not start_date_str or not end_date_str:
        return JsonResponse({'error': 'Rentang tanggal harus diisi'}, status=400)

    from akademik.models import PresensiSekolah, Siswa, Kelas
    
    # Ambil semua siswa yang relevan berdasarkan filter
    siswa_qs = Siswa.objects.select_related('kelas').all()
    if scope == 'tingkatan' and scope_value:
        # Asumsi nama kelas mengandung tingkatan (misal: "Kelas 7A", "Kelas 8B")
        siswa_qs = siswa_qs.filter(kelas__nama_kelas__startswith=f'Kelas {scope_value}')
    elif scope == 'kelas' and scope_value:
        siswa_qs = siswa_qs.filter(kelas_id=scope_value)

    # Ambil semua data presensi dalam rentang tanggal untuk siswa yang relevan
    presensi_data = PresensiSekolah.objects.filter(
        tanggal__range=[start_date_str, end_date_str],
        siswa__in=siswa_qs
    )

    # Lakukan agregasi di Python untuk fleksibilitas
    rekap_per_siswa = {}
    
    # Inisialisasi setiap siswa
    for siswa in siswa_qs:
        rekap_per_siswa[siswa.id] = {
            'id': siswa.id,
            'nisn': siswa.nisn,
            'nama': siswa.nama,
            'kelas': siswa.kelas.nama_kelas,
            'hadir': 0,
            'tidak_hadir': 0,
            'total_hari': 0
        }

    # Proses data presensi
    for presensi in presensi_data:
        if presensi.siswa_id in rekap_per_siswa:
            siswa_rekap = rekap_per_siswa[presensi.siswa_id]
            siswa_rekap['total_hari'] += 1
            if presensi.status == 'H':
                siswa_rekap['hadir'] += 1
            else: # S, I, A
                siswa_rekap['tidak_hadir'] += 1
    
    # Ubah format ke list dan hitung persentase
    hasil_akhir = []
    for siswa_id, data in rekap_per_siswa.items():
        if data['total_hari'] > 0:
            persentase = (data['hadir'] / data['total_hari']) * 100
            data['persentase'] = round(persentase)
        else:
            data['persentase'] = 0 # Jika tidak ada data sama sekali
        
        # Ganti nama field agar cocok dengan frontend
        data['hari_efektif'] = data.pop('total_hari')
        
        hasil_akhir.append(data)
        
    return JsonResponse({ 'rekap': sorted(hasil_akhir, key=lambda x: (x['kelas'], x['nama'])) })

def export_rekap_excel(request):
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    scope = request.GET.get('scope')
    scope_value = request.GET.get('scope_value')

    if not start_date_str or not end_date_str:
        return HttpResponse("Rentang tanggal harus diisi", status=400)

    from akademik.models import PresensiSekolah, Siswa
    
    siswa_qs = Siswa.objects.select_related('kelas').all()
    if scope == 'tingkatan' and scope_value:
        siswa_qs = siswa_qs.filter(kelas__nama_kelas__icontains=scope_value)
    elif scope == 'kelas' and scope_value:
        siswa_qs = siswa_qs.filter(kelas_id=scope_value)

    start_date = timezone.datetime.strptime(start_date_str, '%Y-%m-%d').date()
    end_date = timezone.datetime.strptime(end_date_str, '%Y-%m-%d').date()
    total_hari_efektif = (end_date - start_date).days + 1

    presensi_data = PresensiSekolah.objects.filter(
        tanggal__range=[start_date, end_date],
        siswa__in=siswa_qs
    ).values('siswa_id').annotate(hadir_count=Count('pk', filter=Q(status='H')))
    
    counts_map = {item['siswa_id']: item['hadir_count'] for item in presensi_data}

    # Buat file Excel di memori
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Rekap Kehadiran'

    # Buat Header
    headers = ["No", "NISN", "Nama Siswa", "Kelas", "Hari Efektif", "Hadir", "Tidak Hadir", "Persentase (%)"]
    sheet.append(headers)

    # Isi data
    row_num = 1
    for siswa in sorted(list(siswa_qs), key=lambda s: (s.kelas.nama_kelas if s.kelas else '', s.nama)):
        row_num += 1
        hadir = counts_map.get(siswa.id, 0)
        tidak_hadir = total_hari_efektif - hadir
        persentase = (hadir / total_hari_efektif * 100) if total_hari_efektif > 0 else 0
        
        row_data = [
            row_num - 1,
            siswa.nisn,
            siswa.nama,
            siswa.kelas.nama_kelas if siswa.kelas else '-',
            total_hari_efektif,
            hadir,
            tidak_hadir,
            round(persentase)
        ]
        sheet.append(row_data)

    # Atur lebar kolom otomatis
    for col_idx, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_idx)
        sheet.column_dimensions[column_letter].best_fit = True

    # Siapkan respons untuk mengunduh file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="rekap_kehadiran_{start_date_str}_sd_{end_date_str}.xlsx"'
    workbook.save(response)
    
    return response


def get_siswa_by_kelas(request):
    kelas_id = request.GET.get('kelas_id')
    # PERBAIKAN: Menggunakan 'nisn' sesuai field di database kamu
    siswas = Siswa.objects.filter(kelas_id=kelas_id).values('id', 'nama', 'nisn') 
    return JsonResponse(list(siswas), safe=False)

@transaction.atomic
def kenaikan_kelas(request):
    # 1. Ambil semua kelas dan siapkan Lookup
    all_kelas = list(Kelas.objects.all().order_by('nama_kelas'))
    kelas_lookup = {} # Key: (level, suffix), Value: Kelas Object
    
    max_level = -1
    for k in all_kelas:
        match = re.match(r'^(\d+)(.*)', k.nama_kelas)
        if match:
            val = int(match.group(1))
            suffix = match.group(2).strip()
            if val > max_level: max_level = val
            kelas_lookup[(val, suffix)] = k

    # 2. Logic POST (Proses Simpan)
    if request.method == 'POST':
        # Ambil ID siswa yang dicentang (Artinya: TINGGAL KELAS)
        ids_tinggal = request.POST.getlist('siswa_tinggal')
        
        count_sukses = 0
        
        # Ambil semua siswa untuk dicek kelayakannya naik
        # (Kita perlu hitung ulang targetnya di sini karena form tidak mengirim target kelas)
        semua_siswa = Siswa.objects.select_related('kelas').all()
        
        for siswa in semua_siswa:
            # Jika siswa dicentang (ada di list tinggal), SKIP (jangan ubah kelasnya)
            if str(siswa.id) in ids_tinggal:
                continue
            
            if not siswa.kelas: continue
            
            # Cek apakah valid untuk naik kelas
            match = re.match(r'^(\d+)(.*)', siswa.kelas.nama_kelas)
            if match:
                level_asal = int(match.group(1))
                suffix_asal = match.group(2).strip()
                
                # Jangan proses jika sudah tingkat akhir (harusnya lewat menu kelulusan)
                if level_asal >= max_level:
                    continue
                
                # Cari kelas tujuan (Level + 1)
                target_kelas = kelas_lookup.get((level_asal + 1, suffix_asal))
                
                if target_kelas:
                    # EKSEKUSI PINDAH KELAS
                    siswa.kelas = target_kelas
                    siswa.save()
                    count_sukses += 1
        
        messages.success(request, f'Sukses! {count_sukses} siswa berhasil dinaikkan kelasnya.')
        return redirect('kenaikan_kelas')

    # 3. Logic GET (Tampilan)
    daftar_promosi = [] 
    siswa_qs = Siswa.objects.select_related('kelas').all().order_by('kelas__nama_kelas', 'nama')
    
    for siswa in siswa_qs:
        if not siswa.kelas: continue
            
        k_asal = siswa.kelas
        match = re.match(r'^(\d+)(.*)', k_asal.nama_kelas)
        
        status_teks = ""
        target_nama = "-"
        css_class = ""
        can_promote = False
        is_tingkat_akhir = False
        
        if match:
            level_asal = int(match.group(1))
            suffix_asal = match.group(2).strip()
            
            if level_asal == max_level:
                status_teks = f"Tingkat Akhir (Kls {level_asal})"
                css_class = "text-orange-500 font-bold"
                is_tingkat_akhir = True
            else:
                target_kelas = kelas_lookup.get((level_asal + 1, suffix_asal))
                if target_kelas:
                    target_nama = target_kelas.nama_kelas
                    can_promote = True
                else:
                    status_teks = f"Error: Kls {level_asal+1}{suffix_asal} tdk ada"
                    css_class = "text-red-500 font-bold"
        else:
            status_teks = "Format Nama Kelas Invalid"
            css_class = "text-red-500"

        daftar_promosi.append({
            'siswa': siswa,
            'status_teks': status_teks,
            'target_nama': target_nama,
            'css': css_class,
            'can_promote': can_promote,
            'is_tingkat_akhir': is_tingkat_akhir
        })

    return render(request, 'administrasi/siswa/kenaikan_kelas.html', {
        'daftar_promosi': daftar_promosi,
        'max_level': max_level
    })

@transaction.atomic
def kelulusan_siswa(request):
    # 1. Ambil semua kelas
    all_kelas = Kelas.objects.all()
    
    # 2. Logika mencari Tingkat Tertinggi (Angka terbesar di depan nama kelas)
    max_level = -1
    
    # Regex: ^(\d+) artinya cari angka di awal string
    for k in all_kelas:
        match = re.search(r'^(\d+)', k.nama_kelas)
        if match:
            val = int(match.group(1))
            if val > max_level:
                max_level = val
    
    # 3. Filter kelas yang angkanya sama dengan max_level
    kelas_akhir_ids = []
    nama_kelas_akhir = []
    
    if max_level != -1:
        for k in all_kelas:
            match = re.search(r'^(\d+)', k.nama_kelas)
            if match and int(match.group(1)) == max_level:
                kelas_akhir_ids.append(k.id)
                nama_kelas_akhir.append(k.nama_kelas)
    
    # 4. Ambil siswa hanya dari kelas tingkat akhir tersebut
    # Urutkan per kelas biar rapi
    siswa_akhir = Siswa.objects.filter(kelas_id__in=kelas_akhir_ids).select_related('kelas').order_by('kelas__nama_kelas', 'nama')

    if request.method == 'POST':
        # Siswa yang dicentang adalah yang DITAHAN (TIDAK LULUS)
        siswa_tahan_ids = request.POST.getlist('siswa_tahan')
        
        # Siswa yang akan dihapus adalah: Siswa di tingkat akhir KECUALI yang ditahan
        siswa_hapus = siswa_akhir.exclude(id__in=siswa_tahan_ids)
        
        jumlah_lulus = siswa_hapus.count()
        jumlah_tahan = len(siswa_tahan_ids)
        
        # Eksekusi Hapus
        siswa_hapus.delete()
        
        messages.success(request, f'Proses Selesai! {jumlah_lulus} siswa tingkat akhir telah diluluskan (dihapus), {jumlah_tahan} siswa ditahan.')
        return redirect('siswa')

    return render(request, 'administrasi/siswa/kelulusan_siswa.html', {
        'siswa_akhir': siswa_akhir,
        'tingkat': max_level if max_level != -1 else "Tidak Terdeteksi",
        'daftar_kelas_akhir': ", ".join(nama_kelas_akhir)
    })