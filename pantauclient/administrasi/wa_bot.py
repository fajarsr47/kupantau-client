import os
import time
import subprocess
import openpyxl
import threading
import queue
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.conf import settings
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
# --- PERBAIKAN: Menambahkan import By di sini ---
from selenium.webdriver.common.by import By 
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from akademik.models import PresensiSekolah, Kelas

# ==========================================
# BAGIAN 1: KONFIGURASI ANTREAN (QUEUE)
# ==========================================
wa_task_queue = queue.Queue()

def worker():
    print("🤖 [SYSTEM] WA Bot Worker Berjalan... Menunggu tugas.")
    while True:
        item = wa_task_queue.get()
        if item is None:
            break
        
        try:
            file_path, nama_grup, caption = item
            print(f"\n🤖 [WORKER] Memproses antrean: Kirim ke '{nama_grup}'")
            kirim_laporan_wa_otomatis(file_path, nama_grup, caption)
        except Exception as e:
            print(f"❌ [WORKER] Error processing task: {e}")
        finally:
            wa_task_queue.task_done()
            print(f"✅ [WORKER] Tugas selesai. Sisa antrean: {wa_task_queue.qsize()}")

threading.Thread(target=worker, daemon=True).start()


# ==========================================
# BAGIAN 2: GENERATOR EXCEL
# ==========================================
def generate_laporan_excel(kelas_id, tanggal):
    try:
        kelas = Kelas.objects.get(id=kelas_id)
        presensi_list = PresensiSekolah.objects.filter(
            siswa__kelas_id=kelas_id,
            tanggal=tanggal
        ).select_related('siswa').order_by('siswa__nama')

        if not presensi_list.exists():
            return None, None, None, None

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Presensi {kelas.nama_kelas}"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F46E5")
        border_style = Border(
            left=Side(style='thin'), right=Side(style='thin'), 
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        headers = ["No", "NISN", "Nama Siswa", "Status", "Waktu Input"]
        ws.append(headers)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        stats = {'H': 0, 'I': 0, 'S': 0, 'A': 0, 'T': 0}

        for idx, p in enumerate(presensi_list, 1):
            waktu = p.waktu.strftime("%H:%M") if p.waktu else "-"
            status_ket = p.get_status_display()
            if p.status in stats: stats[p.status] += 1
            
            row_data = [idx, p.siswa.nisn, p.siswa.nama, status_ket, waktu]
            ws.append(row_data)
            
            for col_idx in range(1, 6):
                ws.cell(row=idx+1, column=col_idx).border = border_style

        ws.append([])
        ws.append(["Ringkasan Kehadiran:"])
        ws.append([f"Hadir: {stats['H']}"])
        ws.append([f"Izin: {stats['I']}"])
        ws.append([f"Sakit: {stats['S']}"])
        ws.append([f"Alpha: {stats['A']}"])

        tgl_str = tanggal.strftime("%Y-%m-%d")
        nama_file = f"Laporan_{kelas.nama_kelas.replace(' ', '_')}_{tgl_str}.xlsx"
        
        save_dir = os.path.join(settings.BASE_DIR, 'media', 'temp_reports')
        os.makedirs(save_dir, exist_ok=True)
        
        file_path = os.path.join(save_dir, nama_file)
        wb.save(file_path)
        
        return os.path.abspath(file_path), stats, kelas.nama_grup_wa, kelas.nama_kelas

    except Exception as e:
        print(f"❌ Error Excel: {e}")
        return None, None, None, None


# ==========================================
# BAGIAN 3: HELPER COPY (PowerShell)
# ==========================================
def copy_file_to_clipboard(filepath):
    print(f"📋 [CLIPBOARD] Menyalin file: {filepath}")
    cmd = f'Set-Clipboard -Path "{filepath}"'
    subprocess.run(["powershell", "-Command", cmd], shell=True)


# ==========================================
# BAGIAN 4: SELENIUM BOT (Metode Paste)
# ==========================================
def kirim_laporan_wa_otomatis(file_path, nama_grup, caption):
    print("⚙️  [SELENIUM] Menyiapkan Browser...")
    
    bot_profile_path = os.path.join(settings.BASE_DIR, "session_wa_bot")
    
    options = Options()
    options.add_argument(f"user-data-dir={bot_profile_path}")
    options.add_argument("--remote-debugging-port=9222")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-gpu")
    options.add_argument("--start-maximized")

    driver = None
    try:
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
        driver.get("https://web.whatsapp.com")
        
        wait = WebDriverWait(driver, 120)
        
        print("⏳ [SELENIUM] Menunggu Login WA Web...")
        # Perhatikan penggunaan By.XPATH di sini
        wait.until(EC.presence_of_element_located((By.XPATH, '//div[@id="side"]')))
        print("✅ [SELENIUM] Login Terdeteksi.")

        print(f"🔎 [SELENIUM] Mencari grup: {nama_grup}")
        try:
            search_box = wait.until(EC.element_to_be_clickable((By.XPATH, '//div[@contenteditable="true"][@data-tab="3"]')))
            search_box.clear()
            search_box.click()
            time.sleep(0.5)
            
            search_box.send_keys(nama_grup)
            time.sleep(2.5)
            search_box.send_keys(Keys.ENTER)
        except Exception as e:
            print(f"❌ [SELENIUM] Gagal mencari grup: {e}")
            return False

        print("   -> Menunggu chat terbuka...")
        chat_box = wait.until(EC.presence_of_element_located((By.XPATH, '//div[@contenteditable="true"][@data-tab="10"]')))
        time.sleep(1)

        if os.path.exists(file_path):
            copy_file_to_clipboard(file_path)
            time.sleep(1.5)
            
            print("📋 [SELENIUM] Menempelkan file (CTRL+V)...")
            chat_box.send_keys(Keys.CONTROL, 'v')
        else:
            print(f"❌ [SELENIUM] File tidak ditemukan: {file_path}")
            return False

        print("   -> Menunggu preview file...")
        time.sleep(3) 
        
        if caption:
            print("   -> Mengetik caption...")
            action = webdriver.ActionChains(driver)
            for line in caption.split('\n'):
                action.send_keys(line)
                action.key_down(Keys.SHIFT).send_keys(Keys.ENTER).key_up(Keys.SHIFT)
            action.perform()
            time.sleep(1)

        print("🚀 [SELENIUM] Mengirim...")
        webdriver.ActionChains(driver).send_keys(Keys.ENTER).perform()
        
        print(f"✅ [SELENIUM] SUKSES kirim ke {nama_grup}")
        time.sleep(5)
        
        return True

    except Exception as e:
        print(f"❌ [SELENIUM] Error: {e}")
        return False
    finally:
        if driver:
            print("🏁 [SELENIUM] Menutup browser.")
            driver.quit()


# ==========================================
# BAGIAN 5: FUNGSI PEMICU (INTERFACE)
# ==========================================
def proses_laporan_wa(kelas_id, tanggal, mode="MANUAL"):
    path, stats, grup_wa, nama_kelas = generate_laporan_excel(kelas_id, tanggal)
    
    if path and grup_wa:
        tgl_indo = tanggal.strftime("%d-%m-%Y")
        judul = "LAPORAN PRESENSI HARIAN" if mode == "MANUAL" else "LAPORAN OTOMATIS (SCAN)"
        
        caption = (
            f"*{judul}*\n"
            f"Kelas: {nama_kelas}\n"
            f"Tanggal: {tgl_indo}\n\n"
            f"✅ Hadir: {stats['H']}\n"
            f"ℹ️ Izin: {stats['I']}\n"
            f"🏥 Sakit: {stats['S']}\n"
            f"❌ Alpha: {stats['A']}\n\n"
            f"> _Sent Via Pantauan.my.id_"
        )
        
        wa_task_queue.put((path, grup_wa, caption))
        print(f"📥 [QUEUE] Laporan {nama_kelas} ditambahkan ke antrean.")
        return True
    
    return False